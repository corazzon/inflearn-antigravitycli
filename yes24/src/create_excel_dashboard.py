"""
Yes24 베스트셀러 데이터를 활용하여 엑셀 대시보드를 생성하는 고도화된 스크립트입니다.
이 스크립트는 원천 데이터를 전처리하여 'Data' 시트에 로드하고,
엑셀 공식 수식을 기반으로 Key Metrics와 요약 테이블을 계산하는 'Dashboard' 시트를 구축합니다.
또한, 상위 7개 출판사별 전용 시트와 그 외 나머지 출판사 데이터를 모은 '기타 출판사' 시트를 
각각 개별적으로 구축하여 가독성과 분석 편의성을 극대화합니다.
"""

import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def clean_numeric(val):
    """문자열에서 숫자만 추출하여 실수로 변환합니다."""
    if pd.isna(val):
        return np.nan
    if isinstance(val, str):
        val = val.replace(',', '')
    try:
        return float(val)
    except:
        return np.nan

def extract_point(point_str):
    """포인트 적립 텍스트에서 숫자만 추출합니다."""
    if pd.isna(point_str):
        return 0
    if not isinstance(point_str, str):
        return 0
    nums = re.findall(r'\d+', point_str.replace(',', ''))
    if nums:
        return int(nums[0])
    return 0

def extract_year(date_str):
    """출판 날짜에서 연도를 추출합니다."""
    if pd.isna(date_str) or not isinstance(date_str, str):
        return np.nan
    match = re.search(r'(\d{4})년', date_str)
    return int(match.group(1)) if match else np.nan

def style_data_sheet(ws, fill, font, border):
    """일반 데이터 시트의 헤더 및 데이터 정렬, 서식을 정비하는 헬퍼 함수입니다."""
    ws.views.sheetView[0].showGridLines = True
    
    # 헤더 스타일링
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        
    # 데이터 스타일링 및 정렬
    for row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row, 1):
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            
            # 수치형 정렬 및 포맷팅 (컬럼 이름 또는 데이터 타입 기준)
            # 8: 할인율, 9: 판매가, 10: 정가, 11: 포인트, 12: 판매지수, 13: 리뷰수, 14: 평점
            if col_idx in [2, 7, 8, 11, 12, 13]:  # 순위, 연도, 할인율, 포인트, 판매지수, 리뷰수
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            elif col_idx in [9, 10]:  # 판매가, 정가
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '₩#,##0'
            elif col_idx == 14:  # 평점
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.00'
            elif col_idx in [1, 3, 15]:  # goods_no, goods_type, spring_service
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # 컬럼 너비 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

def main():
    # 데이터 로드
    df_raw = pd.read_csv("yes24/data/yes24_bestsellers.csv")
    
    # 데이터 전처리
    df_clean = pd.DataFrame()
    df_clean['goods_no'] = df_raw['goods_no']
    df_clean['rank'] = df_raw['rank']
    df_clean['goods_type'] = df_raw['goods_type']
    df_clean['goods_name'] = df_raw['goods_name']
    df_clean['author'] = df_raw['author']
    df_clean['publisher'] = df_raw['publisher']
    df_clean['publish_year'] = df_raw['publish_date'].apply(extract_year)
    df_clean['discount_rate'] = df_raw['discount_rate'].fillna(0).astype(int)
    df_clean['sale_price'] = df_raw['sale_price'].apply(clean_numeric)
    df_clean['original_price'] = df_raw['original_price'].apply(clean_numeric)
    df_clean['point'] = df_raw['point'].apply(extract_point)
    df_clean['sale_index'] = df_raw['sale_index'].apply(clean_numeric)
    df_clean['review_count'] = df_raw['review_count'].apply(clean_numeric)
    df_clean['rating'] = df_raw['rating']
    df_clean['spring_service'] = df_raw['spring_service']

    # 워크북 생성
    wb = openpyxl.Workbook()
    
    # 1. Dashboard 시트 생성
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # 2. Data 시트 생성 및 전체 데이터 쓰기
    ws_data = wb.create_sheet(title="Data")
    for r in dataframe_to_rows(df_clean, index=False, header=True):
        ws_data.append(r)
        
    # 스타일 속성 정의
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 전체 데이터 시트 스타일링 적용
    style_data_sheet(ws_data, header_fill, header_font, thin_border)
    
    # 3. 상위 출판사 7곳 추출 및 개별 시트 작성
    # 도서 수가 가장 많은 출판사 상위 7곳
    top_7_publishers = df_clean['publisher'].value_counts().head(7).index.tolist()
    
    for pub in top_7_publishers:
        # 시트명 생성 (슬래시나 특수문자 제거 및 30자 길이 제한 규칙 준수)
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', pub)[:30].strip()
        ws_pub = wb.create_sheet(title=sheet_title)
        
        # 필터링 및 데이터 추가
        df_pub = df_clean[df_clean['publisher'] == pub]
        for r in dataframe_to_rows(df_pub, index=False, header=True):
            ws_pub.append(r)
            
        # 개별 출판사 시트 스타일링 적용
        style_data_sheet(ws_pub, header_fill, header_font, thin_border)
        
    # 4. 나머지 출판사 시트 작성
    ws_etc = wb.create_sheet(title="기타 출판사")
    df_etc = df_clean[~df_clean['publisher'].isin(top_7_publishers)]
    for r in dataframe_to_rows(df_etc, index=False, header=True):
        ws_etc.append(r)
        
    # 기타 출판사 시트 스타일링 적용
    style_data_sheet(ws_etc, header_fill, header_font, thin_border)
        
    # 5. Dashboard 구성 및 스타일링
    # 폰트 정의
    title_font = Font(name="Arial", size=18, bold=True, color="1F4E78")
    section_font = Font(name="Arial", size=13, bold=True, color="000000")
    label_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    card_label_font = Font(name="Arial", size=10, bold=False, color="595959")
    card_val_font = Font(name="Arial", size=16, bold=True, color="1F4E78")
    
    # 색상 채우기 정의
    theme_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

    # A. 타이틀
    ws_dash["B2"] = "YES24 IT/컴퓨터 분야 베스트셀러 분석 대시보드"
    ws_dash["B2"].font = title_font
    
    # B. KPI 카드 (B4:E5 구간)
    # 총 도서 수 카드
    ws_dash["B4"] = "총 베스트셀러 도서 수"
    ws_dash["B4"].font = card_label_font
    ws_dash["B4"].fill = card_fill
    ws_dash["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["B5"] = "=COUNTA(Data!D:D)-1"
    ws_dash["B5"].font = card_val_font
    ws_dash["B5"].fill = card_fill
    ws_dash["B5"].number_format = '#,##0" 권"'
    ws_dash["B5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["B4"].border = thin_border
    ws_dash["B5"].border = thin_border
    
    # 평균 판매가 카드
    ws_dash["C4"] = "도서 평균 판매가"
    ws_dash["C4"].font = card_label_font
    ws_dash["C4"].fill = card_fill
    ws_dash["C4"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["C5"] = "=AVERAGE(Data!I:I)"
    ws_dash["C5"].font = card_val_font
    ws_dash["C5"].fill = card_fill
    ws_dash["C5"].number_format = '₩#,##0'
    ws_dash["C5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["C4"].border = thin_border
    ws_dash["C5"].border = thin_border

    # 평균 판매지수 카드
    ws_dash["D4"] = "도서 평균 판매지수"
    ws_dash["D4"].font = card_label_font
    ws_dash["D4"].fill = card_fill
    ws_dash["D4"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["D5"] = "=AVERAGE(Data!L:L)"
    ws_dash["D5"].font = card_val_font
    ws_dash["D5"].fill = card_fill
    ws_dash["D5"].number_format = '#,##0'
    ws_dash["D5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["D4"].border = thin_border
    ws_dash["D5"].border = thin_border

    # 분철 서비스 비율 카드
    ws_dash["E4"] = "분철 서비스 제공 비율"
    ws_dash["E4"].font = card_label_font
    ws_dash["E4"].fill = card_fill
    ws_dash["E4"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["E5"] = '=COUNTIF(Data!O:O, "Y")/COUNTA(Data!O:O)'
    ws_dash["E5"].font = card_val_font
    ws_dash["E5"].fill = card_fill
    ws_dash["E5"].number_format = '0.0%'
    ws_dash["E5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["E4"].border = thin_border
    ws_dash["E5"].border = thin_border

    # C. 요약 테이블 1 - 상위 출판사별 요약 (점유율 상위 8개사)
    ws_dash["B8"] = "주요 출판사별 베스트셀러 점유 현황"
    ws_dash["B8"].font = section_font
    
    headers_t1 = ["출판사", "도서 수 (권)", "평균 판매지수", "평균 평점"]
    for i, h in enumerate(headers_t1):
        cell = ws_dash.cell(row=9, column=2+i, value=h)
        cell.font = label_font
        cell.fill = theme_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    top_8_publishers = df_clean['publisher'].value_counts().head(8).index.tolist()
    
    for row_idx, pub in enumerate(top_8_publishers):
        r_num = 10 + row_idx
        c_pub = ws_dash.cell(row=r_num, column=2, value=pub)
        c_count = ws_dash.cell(row=r_num, column=3, value=f'=COUNTIF(Data!F:F, "{pub}")')
        c_sale = ws_dash.cell(row=r_num, column=4, value=f'=AVERAGEIF(Data!F:F, "{pub}", Data!L:L)')
        c_rating = ws_dash.cell(row=r_num, column=5, value=f'=AVERAGEIF(Data!F:F, "{pub}", Data!N:N)')
        
        c_count.number_format = '#,##0'
        c_sale.number_format = '#,##0'
        c_rating.number_format = '0.00'
        
        for c in [c_pub, c_count, c_sale, c_rating]:
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
            if c != c_pub:
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
                
    # D. 요약 테이블 2 - 분철 서비스 여부별 비교
    ws_dash["G8"] = "분철 서비스 여부별 비교 분석"
    ws_dash["G8"].font = section_font
    
    headers_t2 = ["분철 가능 여부", "도서 수 (권)", "평균 판매지수"]
    for i, h in enumerate(headers_t2):
        cell = ws_dash.cell(row=9, column=7+i, value=h)
        cell.font = label_font
        cell.fill = theme_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    spring_options = ["Y", "N"]
    for row_idx, opt in enumerate(spring_options):
        r_num = 10 + row_idx
        c_opt = ws_dash.cell(row=r_num, column=7, value="분철 서비스 지원 (Y)" if opt == "Y" else "분철 서비스 미지원 (N)")
        c_count = ws_dash.cell(row=r_num, column=8, value=f'=COUNTIF(Data!O:O, "{opt}")')
        c_sale = ws_dash.cell(row=r_num, column=9, value=f'=AVERAGEIF(Data!O:O, "{opt}", Data!L:L)')
        
        c_count.number_format = '#,##0'
        c_sale.number_format = '#,##0'
        
        for c in [c_opt, c_count, c_sale]:
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
            if c != c_opt:
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # E. 요약 테이블 3 - 가격 구간별 분포
    ws_dash["G13"] = "가격대 구간별 도서 분포"
    ws_dash["G13"].font = section_font
    
    headers_t3 = ["가격대 구간", "도서 수 (권)", "평균 판매지수"]
    for i, h in enumerate(headers_t3):
        cell = ws_dash.cell(row=14, column=7+i, value=h)
        cell.font = label_font
        cell.fill = theme_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    price_ranges = [
        ("1.5만 원 미만", "<15000"),
        ("1.5만 원 이상 ~ 2.5만 원 미만", ">=15000", "<25000"),
        ("2.5만 원 이상 ~ 3.5만 원 미만", ">=25000", "<35000"),
        ("3.5만 원 이상", ">=35000")
    ]
    
    for row_idx, pr in enumerate(price_ranges):
        r_num = 15 + row_idx
        c_range = ws_dash.cell(row=r_num, column=7, value=pr[0])
        
        if len(pr) == 2:
            c_count = ws_dash.cell(row=r_num, column=8, value=f'=COUNTIF(Data!I:I, "{pr[1]}")')
            c_sale = ws_dash.cell(row=r_num, column=9, value=f'=AVERAGEIF(Data!I:I, "{pr[1]}", Data!L:L)')
        else:
            c_count = ws_dash.cell(row=r_num, column=8, value=f'=COUNTIFS(Data!I:I, "{pr[1]}", Data!I:I, "{pr[2]}")')
            c_sale = ws_dash.cell(row=r_num, column=9, value=f'=AVERAGEIFS(Data!L:L, Data!I:I, "{pr[1]}", Data!I:I, "{pr[2]}")')
            
        c_count.number_format = '#,##0'
        c_sale.number_format = '#,##0'
        
        for c in [c_range, c_count, c_sale]:
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
            if c != c_range:
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
                
    # F. 대시보드 내 차트 추가 (출판사별 점유 차트)
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "주요 출판사별 베스트셀러 등록 도서 수"
    chart1.y_axis.title = "출판사"
    chart1.x_axis.title = "도서 수 (권)"
    
    data_ref = Reference(ws_dash, min_col=3, min_row=9, max_row=17)
    cats_ref = Reference(ws_dash, min_col=2, min_row=10, max_row=17)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.legend = None
    chart1.width = 15
    chart1.height = 10
    
    ws_dash.add_chart(chart1, "B19")
    
    # G. 대시보드 내 차트 추가 2 (분철 여부별 판매지수 차트)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.title = "분철 여부별 평균 판매지수 비교"
    chart2.y_axis.title = "평균 판매지수"
    chart2.x_axis.title = "분철 여부"
    
    data_ref2 = Reference(ws_dash, min_col=9, min_row=9, max_row=11)
    cats_ref2 = Reference(ws_dash, min_col=7, min_row=10, max_row=11)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    chart2.legend = None
    chart2.width = 13
    chart2.height = 10
    
    ws_dash.add_chart(chart2, "G20")

    # 대시보드 컬럼 너비 정밀 세팅
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 25
    ws_dash.column_dimensions["C"].width = 16
    ws_dash.column_dimensions["D"].width = 18
    ws_dash.column_dimensions["E"].width = 14
    ws_dash.column_dimensions["F"].width = 3
    ws_dash.column_dimensions["G"].width = 28
    ws_dash.column_dimensions["H"].width = 16
    ws_dash.column_dimensions["I"].width = 18

    # 저장
    output_file = "yes24/docs/Yes24_Bestseller_Dashboard.xlsx"
    wb.save(output_file)
    print(f"엑셀 대시보드 파일 생성 및 출판사 시트 분할 완료: {output_file}")

if __name__ == "__main__":
    main()
